"""Round-trip the report content through Excel so a human can edit it.

The generated .docx is a rendered artifact -- editing it means fighting page
layout, photo grids and sizing. This gives the reviewer a plain spreadsheet
instead: one row per item, edit the words, tick a box to drop an item, type a
number to reorder, add a row to insert one. Then re-render.

    export : master_report_items.json  ->  Report-Review.xlsx
    import : Report-Review.xlsx        ->  master_report_items.json (updated)

Nothing but the reviewer-owned columns is read back on import, so photo paths,
sheet clips and dimensions cannot be corrupted by editing the sheet.

Usage:
  python review_sheet.py export <build_dir> [-o Report-Review.xlsx]
  python review_sheet.py import <build_dir> <Report-Review.xlsx>
"""

from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BLUE, DARK, YELLOW, GREY = "666F89", "58595B", "FFF2CC", "F2F2F2"
FONT = "Arial"

# (header, source key, editable?, width)
COLUMNS = [
    ("Include?",        "_include",       True,  10),
    ("Order",           "_order",         True,   8),
    ("PlanGrid ref",    "number",         False, 12),
    ("Sheet",           "sheet_name",     False, 12),
    ("Location",        "location",       True,  34),
    ("Description",     "description",    True,  70),
    ("Reviewer flag",   "reviewer_flag",  True,  34),
    ("Confidence / Source", "confidence", False, 26),
    ("Photos",          "_photos",        False,  8),
]
EDITABLE = [k for _, k, e, _ in COLUMNS if e]


def export(build: Path, out: Path) -> None:
    master = json.loads((build / "master_report_items.json").read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    ws.append([h for h, _, _, _ in COLUMNS])
    for i, (h, _, editable, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for n, item in enumerate(master, start=1):
        row = []
        for _, key, _, _ in COLUMNS:
            if key == "_include":
                row.append("Y")
            elif key == "_order":
                row.append(n * 10)          # gaps, so a row can be slotted between
            elif key == "_photos":
                row.append(len(item.get("photo_paths") or []))
            else:
                row.append(item.get(key, ""))
        ws.append(row)

    for i, (_, key, editable, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=i)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if editable:
                c.fill = PatternFill("solid", fgColor=YELLOW)
            else:
                c.fill = PatternFill("solid", fgColor=GREY)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 46

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    notes = wb.create_sheet("How to use")
    for i, line in enumerate([
        "Editing this report",
        "",
        "Yellow cells are yours to edit. Grey cells are generated - changing them does nothing.",
        "",
        "Drop an item          set Include? to N. Remaining items renumber themselves.",
        "Reorder               change Order. Values are spaced by 10 so you can slot one between.",
        "Reword                edit Location, Description or Reviewer flag.",
        "Add an item           add a row, set Include? = Y and an Order value, leave PlanGrid ref blank.",
        "                      A new row has no photos or sheet clip - add those in Word afterwards.",
        "",
        "Then re-render:   python review_sheet.py import <build_dir> <this file>",
        "                  node gen_report.js <build_dir>",
        "",
        "PlanGrid ref is the permanent link back to PlanGrid and is never renumbered.",
        "The Item number printed in the report is a Word auto-number and will renumber itself.",
    ], start=1):
        c = notes.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT, size=11 if i == 1 else 10,
                      bold=(i == 1), color=BLUE if i == 1 else DARK)
    notes.column_dimensions["A"].width = 110

    wb.save(out)
    print(f"exported {len(master)} items -> {out}")


def do_import(build: Path, sheet: Path) -> None:
    master_path = build / "master_report_items.json"
    master = json.loads(master_path.read_text(encoding="utf-8"))
    by_ref = {str(m.get("number")): m for m in master}

    ws = load_workbook(sheet, data_only=True)["Items"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    kept, dropped, added, edited = [], 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v in (None, "") for v in row):
            continue
        include = str(row[idx["Include?"]] or "Y").strip().upper()
        if include.startswith("N"):
            dropped += 1
            continue
        ref = row[idx["PlanGrid ref"]]
        order = row[idx["Order"]]
        base = by_ref.get(str(ref)) if ref not in (None, "") else None
        if base is None:
            base = {"number": ref if ref not in (None, "") else None,
                    "photo_paths": [], "origin": "manual",
                    "confidence": "Added during review"}
            added += 1
        else:
            before = (base.get("description"), base.get("location"),
                      base.get("reviewer_flag"))
            after = (row[idx["Description"]], row[idx["Location"]],
                     row[idx["Reviewer flag"]])
            if before != after:
                edited += 1
        base["location"] = row[idx["Location"]] or ""
        base["description"] = row[idx["Description"]] or ""
        base["reviewer_flag"] = row[idx["Reviewer flag"]] or ""
        base["_order"] = order if isinstance(order, (int, float)) else 10_000
        kept.append(base)

    kept.sort(key=lambda m: m["_order"])
    for m in kept:
        m.pop("_order", None)

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = master_path.with_name(f"master_report_items.{stamp}.bak.json")
    shutil.copy2(master_path, backup)
    master_path.write_text(json.dumps(kept, indent=1, ensure_ascii=False),
                           encoding="utf-8")
    print(f"kept {len(kept)}  dropped {dropped}  added {added}  reworded {edited}")
    print(f"backup: {backup.name}")
    print(f"updated: {master_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("mode", choices=["export", "import"])
    ap.add_argument("build_dir")
    ap.add_argument("sheet", nargs="?")
    ap.add_argument("-o", "--out", default="Report-Review.xlsx")
    args = ap.parse_args()

    build = Path(args.build_dir)
    if args.mode == "export":
        export(build, Path(args.out))
    else:
        if not args.sheet:
            raise SystemExit("import needs the review .xlsx path")
        do_import(build, Path(args.sheet))


if __name__ == "__main__":
    main()
